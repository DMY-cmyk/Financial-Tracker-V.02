#!/usr/bin/env python3
"""
Extract data from Financial Tracker.xlsx into data/workbook.json.

Usage:
    pip install openpyxl
    python scripts/extract_xlsx.py
    # or specify a custom path:
    python scripts/extract_xlsx.py path/to/spreadsheet.xlsx
"""

import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

# Month name mapping for Indonesian-style month abbreviations
MONTH_NAMES = [
    "JAN", "FEB", "MAR", "APR", "MEI", "JUN",
    "JUL", "AGU", "SEP", "OKT", "NOV", "DES"
]


def col_letter_to_index(letter):
    """Convert column letter(s) to 0-based index (A=0, B=1, ..., Z=25, AA=26)."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1


def extract_sheet(ws, sheet_id, sheet_name):
    """Extract a single worksheet into a dict."""
    cells = {}
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            value = cell.value
            if value is None:
                continue

            r = cell.row - 1  # 0-based
            c = cell.column - 1  # 0-based
            key = f"{r}:{c}"

            raw = str(value) if not isinstance(value, str) else value

            # Preserve formulas if available
            formula = ""
            if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value

            cells[key] = {
                "raw": raw,
                "formula": formula
            }

    return {
        "id": sheet_id,
        "name": sheet_name,
        "rowCount": max_row,
        "colCount": max_col,
        "cells": cells
    }


def extract_workbook(xlsx_path):
    """Extract all sheets from an xlsx file."""
    wb = load_workbook(xlsx_path, data_only=False)
    sheets = []

    for idx, ws_name in enumerate(wb.sheetnames):
        ws = wb[ws_name]
        sheet_id = f"sheet-{idx + 1}"

        # Try to match month names for the sheet name
        name = ws_name.strip()
        sheets.append(extract_sheet(ws, sheet_id, name))

    wb.close()
    return {"sheets": sheets}


def main():
    # Determine project root (parent of scripts/)
    project_root = Path(__file__).resolve().parent.parent

    # Default xlsx path
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        xlsx_path = project_root / "Financial Tracker.xlsx"

    if not xlsx_path.exists():
        print(f"Error: Spreadsheet not found at {xlsx_path}")
        print("Place 'Financial Tracker.xlsx' in the project root, or pass a path as argument.")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    workbook_data = extract_workbook(xlsx_path)

    output_dir = project_root / "data"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / "workbook.json"

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(workbook_data, f, ensure_ascii=False, indent=2)

    total_cells = sum(len(s["cells"]) for s in workbook_data["sheets"])
    print(f"Extracted {len(workbook_data['sheets'])} sheets, {total_cells} cells")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
